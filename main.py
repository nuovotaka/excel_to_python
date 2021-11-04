from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment
import PySimpleGUI as sg
import sys

# 関数定義


def del_val_line(ws, print_sheet_title, min_row, init_max_row, init_max_column):
    # 値と罫線の削除
    border_clear = Border(top=None, bottom=None, left=None, right=None)

    if init_max_row > print_sheet_title:
        for r in range(min_row, init_max_row+1):
            ws.row_dimensions[r].height = 13.5  # ２行目以降の高さを13.5pointに設定
            for col in range(1, init_max_column+1):
                ws.cell(r, col).value = None
                ws.cell(r, col).border = border_clear


def disp_line(ws, print_sheet_title, print_end, col_max):

    # 罫線の初期設定
    side_hair = Side(style='hair', color='000000')
    side_thin = Side(style='thin', color='000000')
    border_hair = Border(top=side_hair, bottom=side_hair,
                         left=side_hair, right=side_hair)
    border = Border(top=side_thin, bottom=side_thin,
                    left=side_hair, right=side_hair)

    for r in range(print_sheet_title+1, print_end+1):
        ws.row_dimensions[r].height = 22.5  # 行の高さ調節
        for col in range(1, col_max):
            ws.cell(row=r, column=col).border = border_hair

    for r in range(print_sheet_title, print_sheet_title+1):
        ws.row_dimensions[r].height = 24
        for col in range(1, col_max):
            ws.cell(row=r, column=col).border = border

    for r in range(print_end, print_end+1):
        for col in range(1, col_max):
            ws.cell(row=r, column=col).border = border


# 処理開始
# 初期化処理
# 入力データーの列番号の指定 見積内訳(入力)
input_array = ("", 2, 3, 4, 5, 6, 8, 11)

# 印刷データの列番号
# 縦見積のデーターの指定 縦見積(印刷)
output_tate_array = ("", 2, 3, 4, 5, 6, 8)

# 内訳のデーターの指定 内訳(印刷)
output_yoko_array = ("", 2, 3, 4, 5, 6, 7, 8)

# 集計部分の項目名 初期化
out_sum_item = ("", "", '諸経費', '小　計', '消費税', '合　計')

# 印刷シートのタイトル行が何行目か？(定数の初期化)
OUT_PRINT_ST = 1
OUT_PRINT_ST_TATE = 11

# ファイル名を指定して開く
f = sg.popup_get_file('Excel ファイルを選択', file_types=(('Excel file', '*.xlsx')))

if f == None or f == "":
    sg.popup('最初からやり直してください！')
    sys.exit()

# 印刷する最終行を入力
value = sg.popup_get_text('印刷させる最後の項目番号は何番ですか？', '印刷最終行入力')

if not (value == None or value == ""):
    out_data_cnt = int(value)
else:
    sg.popup('最初からやり直してください！')
    sys.exit()

wb = load_workbook(
    filename=f, data_only=True)

# ワークシートの設定
ws1 = wb["表紙(印刷)"]
ws2 = wb["内訳(印刷)"]
ws3 = wb["見積内訳(入力)"]
ws4 = wb["縦見積り用（印刷）"]

# ワークシートの保護を解除
ws1.protection.disable()
ws2.protection.disable()
ws3.protection.disable()
ws4.protection.disable()

# 片付け清掃、諸経費の表示の有無(フラグの初期化)
disp_sw_flag = ws3.cell(row=1, column=10)  # J1

# 内訳印刷用シートの初期化
initial_rows = int(ws2.max_row)
initial_columns = int(ws2.max_column)
minimum_row = OUT_PRINT_ST + 1

# 値と罫線の削除

del_val_line(ws2, OUT_PRINT_ST, minimum_row, initial_rows, initial_columns)

# 縦見積 印刷用シートの初期化
initial_rows = int(ws4.max_row)
initial_columns = int(ws4.max_column)
minimum_row = OUT_PRINT_ST_TATE + 1

# 値と罫線の削除

del_val_line(ws4, OUT_PRINT_ST_TATE, minimum_row,
             initial_rows, initial_columns)

# ワークシートを入力シートへ移動

# 内訳印刷シートへコピー
# 入力シート ＝＞ 印刷シート へ値のコピー＆ペースト
input_sheet_row = 8  # 8行目, セルではB8がコピー元の最初の位置
output_sheet_row = OUT_PRINT_ST + 1  # タイトル行＋データペースト先の指定

for r in range(1, out_data_cnt+1):
    for col in range(1, 8):
        ws2.cell(output_sheet_row, output_yoko_array[col]).value = ws3.cell(
            input_sheet_row, input_array[col]).value
        if col == 6 or col == 7:
            ws2.cell(output_sheet_row,
                     output_yoko_array[col]).number_format = "###,###,##0"
    input_sheet_row = input_sheet_row + 1
    output_sheet_row = output_sheet_row + 1

# 内訳集計部分を追加(諸経費、小計、消費税、合計)
output_sheet_row = out_data_cnt+OUT_PRINT_ST+3

for i in range(2, 6):
    if not (disp_sw_flag == 1 and i == 2):
        ws2.cell(output_sheet_row, 2).value = out_sum_item[i]
        ws2.cell(output_sheet_row, 7).value = ws3.cell(i, 3).value
        ws2.cell(output_sheet_row, 7).number_format = "###,###,##0"
    output_sheet_row = output_sheet_row + 1  # 諸経費のタイトルと金額を表示しない

# 縦見積 印刷シートへのコピー
# 入力シート ＝＞ 印刷シート へ値のコピー＆ペースト
input_sheet_row = 8  # 8行目, セルではB8がコピー元の最初の位置
output_sheet_row = OUT_PRINT_ST_TATE + 1  # タイトル行＋データペースト先の指定

for r in range(1, out_data_cnt+1):
    for col in range(1, 7):
        ws4.cell(output_sheet_row, col).value = ws3.cell(
            input_sheet_row, output_tate_array[col]).value
        if col == 5 or col == 6:
            ws4.cell(output_sheet_row, col).number_format = "###,###,##0"
    input_sheet_row = input_sheet_row + 1
    output_sheet_row = output_sheet_row + 1

# 縦見積 集計部分を追加(諸経費、小計、消費税、合計)
output_sheet_row = out_data_cnt+OUT_PRINT_ST_TATE+3

for i in range(2, 6):
    if not (disp_sw_flag == 1 and i == 2):
        ws4.cell(output_sheet_row, 1).value = out_sum_item[i]
        ws4.cell(output_sheet_row, 6).value = ws3.cell(i, 3).value
        ws4.cell(output_sheet_row, 6).number_format = "###,###,##0"
    output_sheet_row = output_sheet_row + 1  # 諸経費のタイトルと金額を表示しない

# 罫線の描画範囲を指定

# 内訳見積
print_end = out_data_cnt + OUT_PRINT_ST + 6
col_max = 9

disp_line(ws2, OUT_PRINT_ST, print_end, col_max)

# 縦見積
print_end = out_data_cnt + OUT_PRINT_ST_TATE + 6
col_max = 7

disp_line(ws4, OUT_PRINT_ST_TATE, print_end, col_max)

# 表紙(印刷へ合計金額と消費税をコピペ
font_sum = Font(name='MS　Pゴシック', size=28, bold=True, color='000000')

ws1['E14'].value = ws3['C5'].value
# E14~F14のセルを結合
ws1.merge_cells(range_string="E14:F14")
ws1['E14'].number_format = "￥ ###,###,##0.-"
ws1['E14'].font = font_sum

# E14セルに［右揃え］を設定
ws1["E14"].alignment = Alignment(horizontal="right", vertical="center")

# H14~J14のセルを結合
font_zei = Font(name='MS　Pゴシック', size=11, color='000000')

ws1['H14'].value = ws3['C4'].value
ws1.merge_cells(range_string="H14:J14")
ws1['H14'].number_format = "(内消費税分　 ##,###,##0 )"
ws1['H14'].font = font_zei

ws1["H14"].alignment = Alignment(horizontal="left", vertical="center")

# 縦見積へ金額、消費税、合計金額をコピペ
font_tate1 = Font(name='MS　Pゴシック', size=12, bold=False, color='000000')

ws4['B6'].number_format = "￥ ###,###,##0"
ws4['B6'].value = ws3['C3'].value
ws4['B6'].font = font_tate1

ws4["B6"].alignment = Alignment(horizontal="right", vertical="center")

ws4['B7'].number_format = "￥ ###,###,##0"
ws4['B7'].value = ws3['C4'].value
ws4['B7'].font = font_tate1

ws4["B7"].alignment = Alignment(horizontal="right", vertical="center")

font_tate2 = Font(name='MS　Pゴシック', size=14, bold=True, color='000000')

ws4['B8'].number_format = "￥ ###,###,##0"
ws4['B8'].value = ws3['C5'].value
ws4['B8'].font = font_tate2

ws4["B8"].alignment = Alignment(horizontal="right", vertical="center")

# ワークシートを保護
ws1.protection.enable()
ws2.protection.enable()
ws3.protection.enable()
ws4.protection.enable()

# 保存＆終了　処理
wb.save(f)
wb.close()

sg.popup('入力シートから印刷シートへコピーできました')
